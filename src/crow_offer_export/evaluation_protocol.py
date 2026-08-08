from __future__ import annotations

from io import BytesIO

from crow_pressure_test.workflow import PressureTestEvaluation

from .offer_workbook import _require_openpyxl, _styles


def build_evaluation_protocol_workbook(result: PressureTestEvaluation) -> bytes:
    """Build an XLSX protocol from one evaluated pressure test.

    The workbook records the exact values already evaluated by
    ``crow_pressure_test.workflow``. It does not recalculate q_max.
    """
    _require_openpyxl()
    import openpyxl

    styles = _styles()
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Provningsresultat"
    sheet.column_dimensions["A"].width = 30
    sheet.column_dimensions["B"].width = 28
    sheet["A1"] = "TÄTHETSPROVNING – RESULTAT"
    sheet["A1"].font = styles["h1"]
    sheet["A1"].fill = styles["fill_h"]
    sheet.merge_cells("A1:B1")

    rows = (
        ("Projekt", result.project_id),
        ("Täthetsklass", result.tightness_class.value),
        ("ATC-klass", result.atc_class),
        ("Provtryck (Pa)", str(result.pressure_pa)),
        ("Kanalarea (m²)", str(result.duct_area_m2)),
        ("Tillåtet läckage (l/s)", str(result.allowed_leakage_lps)),
        (
            "Uppmätt läckage (l/s)",
            "" if result.measured_leakage_lps is None else str(result.measured_leakage_lps),
        ),
        ("Resultat", result.status.value.upper()),
        ("Protokollklart", "JA" if result.ready_for_protocol else "NEJ"),
    )
    for row_index, (label, value) in enumerate(rows, start=3):
        sheet.cell(row=row_index, column=1, value=label).font = styles["bb"]
        sheet.cell(row=row_index, column=2, value=value).border = styles["border"]

    provenance = workbook.create_sheet("Proveniens")
    provenance.column_dimensions["A"].width = 24
    provenance.column_dimensions["B"].width = 28
    provenance.column_dimensions["C"].width = 14
    provenance.column_dimensions["D"].width = 54
    provenance.column_dimensions["E"].width = 18
    headers = ("Fält", "Värde", "Origin", "Källreferens", "Bekräftad")
    for column_index, header in enumerate(headers, start=1):
        cell = provenance.cell(row=1, column=column_index, value=header)
        cell.font = styles["th"]
        cell.fill = styles["fill_th"]
        cell.border = styles["border"]
    for row_index, item in enumerate(result.provenance, start=2):
        values = (
            item.field,
            item.value,
            item.origin.value.upper(),
            item.source_ref or "",
            "JA" if item.confirmed else "NEJ",
        )
        for column_index, value in enumerate(values, start=1):
            provenance.cell(row=row_index, column=column_index, value=value).border = styles[
                "border"
            ]

    standards = workbook.create_sheet("Standarder")
    standards.column_dimensions["A"].width = 24
    standards.column_dimensions["B"].width = 72
    for row_index, (standard_id, title) in enumerate(result.standards, start=1):
        standards.cell(row=row_index, column=1, value=standard_id).font = styles["bb"]
        standards.cell(row=row_index, column=2, value=title)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
