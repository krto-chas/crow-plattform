from __future__ import annotations

from collections.abc import Mapping, Sequence
from io import BytesIO
from typing import Any

from .offer_workbook import _require_openpyxl, _styles


def build_evaluation_protocol_workbook(result: Mapping[str, Any]) -> bytes:
    """Build XLSX from an already evaluated pressure-test payload.

    The adapter records values produced by ``crow_pressure_test.workflow`` and
    intentionally performs no leakage-limit recalculation.
    """
    _require_openpyxl()
    import openpyxl

    styles = _styles()
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Provningsresultat"
    sheet.column_dimensions["A"].width = 30
    sheet.column_dimensions["B"].width = 28
    sheet["A1"] = "TÄTHETSPROVNING – RESULTAT"
    sheet["A1"].font = styles["h1"]
    sheet["A1"].fill = styles["fill_h"]
    sheet.merge_cells("A1:B1")

    rows = (
        ("Projekt", str(result.get("project_id", ""))),
        ("Täthetsklass", str(result.get("tightness_class", ""))),
        ("ATC-klass", str(result.get("atc_class", ""))),
        ("Provtryck (Pa)", str(result.get("pressure_pa", ""))),
        ("Kanalarea (m²)", str(result.get("duct_area_m2", ""))),
        ("Tillåtet läckage (l/s)", str(result.get("allowed_leakage_lps", ""))),
        ("Uppmätt läckage (l/s)", _optional_value(result.get("measured_leakage_lps"))),
        ("Resultat", str(result.get("status", "")).upper()),
        ("Protokollklart", "JA" if bool(result.get("ready_for_protocol")) else "NEJ"),
    )
    for row_index, (label, value) in enumerate(rows, start=3):
        sheet.cell(row=row_index, column=1, value=label).font = styles["bb"]
        sheet.cell(row=row_index, column=2, value=value).border = styles["border"]

    provenance = workbook.create_sheet("Proveniens")
    for column, width in zip("ABCDE", (24, 28, 14, 54, 18), strict=False):
        provenance.column_dimensions[column].width = width
    headers = ("Fält", "Värde", "Origin", "Källreferens", "Bekräftad")
    for column_index, header in enumerate(headers, start=1):
        cell = provenance.cell(row=1, column=column_index, value=header)
        cell.font = styles["th"]
        cell.fill = styles["fill_th"]
        cell.border = styles["border"]
    for row_index, item in enumerate(_mapping_items(result.get("provenance")), start=2):
        values = (
            str(item.get("field", "")),
            str(item.get("value", "")),
            str(item.get("origin", "")).upper(),
            _optional_value(item.get("source_ref")),
            "JA" if bool(item.get("confirmed")) else "NEJ",
        )
        for column_index, value in enumerate(values, start=1):
            provenance.cell(row=row_index, column=column_index, value=value).border = styles[
                "border"
            ]

    standards = workbook.create_sheet("Standarder")
    standards.column_dimensions["A"].width = 24
    standards.column_dimensions["B"].width = 72
    for row_index, item in enumerate(_mapping_items(result.get("standards")), start=1):
        standards.cell(row=row_index, column=1, value=str(item.get("id", ""))).font = styles["bb"]
        standards.cell(row=row_index, column=2, value=str(item.get("title", "")))

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _mapping_items(value: object) -> tuple[Mapping[str, Any], ...]:
    if not isinstance(value, Sequence) or isinstance(value, (str, bytes, bytearray)):
        return ()
    return tuple(item for item in value if isinstance(item, Mapping))


def _optional_value(value: object) -> str:
    return "" if value is None else str(value)
