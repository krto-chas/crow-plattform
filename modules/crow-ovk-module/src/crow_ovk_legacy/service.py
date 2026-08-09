from __future__ import annotations

import re
from hashlib import sha256
from io import BytesIO
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from pypdf import PdfReader

from .models import (
    LegacyFact,
    LegacyFactStatus,
    LegacyImportPreview,
    LegacyReviewItem,
    LegacySourceKind,
    LegacySourceRef,
)

_SYSTEM_RE = re.compile(r"\b(?:TA|FA|TF|FF|FTX|FT|FX)\s*[-_]?\s*\d{1,3}\b", re.IGNORECASE)
_DATE_RE = re.compile(r"\b(20\d{2}[-/.](?:0?[1-9]|1[0-2])[-/.](?:0?[1-9]|[12]\d|3[01]))\b")
_APARTMENT_RE = re.compile(r"\b(?:lgh|lägenhet)\s*(?:nr|nummer)?\s*[:#-]?\s*([A-Za-z0-9-]{2,12})\b", re.IGNORECASE)
_MEASURED_RE = re.compile(
    r"(?:uppmätt|mätvärde)\s*[:=]?\s*(-?\d+(?:[.,]\d+)?)\s*(l/s|m3/s|m³/s)",
    re.IGNORECASE,
)
_DESIGNED_RE = re.compile(
    r"(?:projekterat|börvärde|dimensionerande)\s*[:=]?\s*(-?\d+(?:[.,]\d+)?)\s*(l/s|m3/s|m³/s)",
    re.IGNORECASE,
)
_FINDING_RE = re.compile(r"\b(?:anmärkning|brist)\b\s*[:\-]?\s*(.+)", re.IGNORECASE)
_AIRFLOW_RE = re.compile(r"-?\d+(?:[.,]\d+)?\s*(?:l/s|m3/s|m³/s)", re.IGNORECASE)


def preview_legacy_file(project_id: str, filename: str, content: bytes) -> LegacyImportPreview:
    suffix = Path(filename).suffix.lower()
    if suffix == ".pdf":
        kind = LegacySourceKind.PDF
        records = _pdf_records(filename, content)
    elif suffix == ".xlsx":
        kind = LegacySourceKind.XLSX
        records = _xlsx_records(filename, content)
    else:
        raise ValueError(f"Unsupported legacy OVK file type: {suffix or '<none>'}")

    source_digest = sha256(content).hexdigest()
    facts: list[LegacyFact] = []
    review: list[LegacyReviewItem] = []
    for locator, text in records:
        normalized = " ".join(text.split())
        if not normalized:
            continue
        source = LegacySourceRef(
            source_id=f"legacy:{source_digest[:16]}:{locator}",
            filename=filename,
            kind=kind,
            locator=locator,
            sha256=source_digest,
        )
        extracted = _facts_from_text(normalized, source)
        facts.extend(extracted)
        if not extracted and _looks_ovk_relevant(normalized):
            review.append(
                LegacyReviewItem(
                    reason="ovk_relevant_but_unmapped",
                    source_text=normalized,
                    source=source,
                )
            )
        elif _AIRFLOW_RE.search(normalized) and not any(
            fact.field in {"measured_airflow", "designed_airflow"} for fact in extracted
        ):
            review.append(
                LegacyReviewItem(
                    reason="unlabelled_airflow_value",
                    source_text=normalized,
                    source=source,
                )
            )

    return LegacyImportPreview(
        project_id=project_id,
        filename=filename,
        kind=kind,
        source_sha256=source_digest,
        facts=tuple(facts),
        review=tuple(review),
    )


def _pdf_records(filename: str, content: bytes) -> tuple[tuple[str, str], ...]:
    reader = PdfReader(BytesIO(content))
    records: list[tuple[str, str]] = []
    for page_number, page in enumerate(reader.pages, start=1):
        text = page.extract_text() or ""
        for line_number, line in enumerate(text.splitlines(), start=1):
            records.append((f"page:{page_number}:line:{line_number}", line))
    return tuple(records)


def _xlsx_records(filename: str, content: bytes) -> tuple[tuple[str, str], ...]:
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    records: list[tuple[str, str]] = []
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            populated = [cell for cell in row if cell.value not in (None, "")]
            if not populated:
                continue
            text = " | ".join(str(cell.value) for cell in populated)
            locator = f"sheet:{worksheet.title}:cells:{populated[0].coordinate}-{populated[-1].coordinate}"
            records.append((locator, text))
    workbook.close()
    return tuple(records)


def _facts_from_text(text: str, source: LegacySourceRef) -> tuple[LegacyFact, ...]:
    facts: list[LegacyFact] = []
    for raw in _system_ids(text):
        facts.append(_fact("system_id", raw, source))

    date_match = _DATE_RE.search(text)
    if date_match:
        facts.append(_fact("inspection_date", date_match.group(1).replace("/", "-").replace(".", "-"), source))

    apartment = _APARTMENT_RE.search(text)
    if apartment:
        facts.append(_fact("apartment_number", apartment.group(1), source))

    measured = _MEASURED_RE.search(text)
    if measured:
        facts.append(_fact("measured_airflow", f"{_number(measured.group(1))} {_unit(measured.group(2))}", source))

    designed = _DESIGNED_RE.search(text)
    if designed:
        facts.append(_fact("designed_airflow", f"{_number(designed.group(1))} {_unit(designed.group(2))}", source))

    finding = _FINDING_RE.search(text)
    if finding and finding.group(1).strip():
        facts.append(_fact("finding", finding.group(1).strip(), source))

    return tuple(facts)


def _fact(field: str, value: str, source: LegacySourceRef) -> LegacyFact:
    return LegacyFact(
        field=field,
        value=value,
        confidence="explicit",
        status=LegacyFactStatus.EXTRACTED,
        source=source,
    )


def _system_ids(text: str) -> tuple[str, ...]:
    values = {
        re.sub(r"[\s_-]+", "", match.group(0)).upper()
        for match in _SYSTEM_RE.finditer(text)
    }
    return tuple(sorted(values))


def _looks_ovk_relevant(text: str) -> bool:
    lowered = text.lower()
    terms: Iterable[str] = (
        "ovk",
        "ventilation",
        "luftflöde",
        "luftflode",
        "frånluft",
        "tilluft",
        "anmärkning",
        "brist",
        "besiktning",
    )
    return any(term in lowered for term in terms)


def _number(value: str) -> str:
    return value.replace(",", ".")


def _unit(value: str) -> str:
    return "m3/s" if value.lower() in {"m3/s", "m³/s"} else "l/s"
