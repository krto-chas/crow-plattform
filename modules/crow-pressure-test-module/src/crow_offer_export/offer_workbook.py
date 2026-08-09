from __future__ import annotations

from collections.abc import Mapping
from pathlib import Path
from typing import Any

_HEADER_FILL = "1F4E5F"
_TABLE_FILL = "4A7A8C"
_SUM_FILL = "DCE6F1"
_INPUT_FILL = "FFF7CC"


def _require_openpyxl() -> Any:
    try:
        import openpyxl
    except ImportError as error:  # pragma: no cover - miljöberoende
        raise RuntimeError(
            "openpyxl krävs för xlsx-export; installera med extra 'export' "
            "(pip install crow-plattform[export])"
        ) from error
    return openpyxl


def _styles() -> dict[str, Any]:
    from openpyxl.styles import Border, Font, PatternFill, Side

    thin = Side(style="thin", color="B0B0B0")
    return {
        "h1": Font(name="Arial", size=13, bold=True, color="FFFFFF"),
        "th": Font(name="Arial", size=9, bold=True, color="FFFFFF"),
        "b": Font(name="Arial", size=10),
        "bb": Font(name="Arial", size=10, bold=True),
        "note": Font(name="Arial", size=8, italic=True, color="666666"),
        "fill_h": PatternFill("solid", fgColor=_HEADER_FILL),
        "fill_th": PatternFill("solid", fgColor=_TABLE_FILL),
        "fill_sum": PatternFill("solid", fgColor=_SUM_FILL),
        "fill_in": PatternFill("solid", fgColor=_INPUT_FILL),
        "border": Border(left=thin, right=thin, top=thin, bottom=thin),
    }


def write_offer_workbook(
    offer: Mapping[str, Any],
    service_quantities: Mapping[str, Any],
    output_path: Path,
    project_title: str,
) -> Path:
    """Skriver offertkalkylen som xlsx: sammanställning + flik per trapphus.

    Presentationsadapter enligt ADR-0004: all sanning ligger i offert- och
    mängdpayloaderna; arbetsboken är en vy och innehåller inga egna formler
    för beloppen.
    """
    openpyxl = _require_openpyxl()
    styles = _styles()
    money = '#,##0.00" kr"'
    workbook = openpyxl.Workbook()

    summary = workbook.active
    summary.title = "Sammanställning"
    for column, width in zip("ABCDEF", (34, 14, 10, 14, 16, 30), strict=False):
        summary.column_dimensions[column].width = width
    summary["A1"] = f"PROVTRYCKNING – OFFERTKALKYL: {project_title}"
    summary["A1"].font = styles["h1"]
    summary["A1"].fill = styles["fill_h"]
    summary.merge_cells("A1:F1")
    summary["A2"] = (
        f"Prisbok: {offer['price_book_id']}  |  Valuta: {offer['currency']}  |  "
        f"Schema: {offer['schema_version']}"
    )
    summary["A2"].font = styles["note"]

    headers = ("Post", "Trapphus/del", "Antal", "Enhet", "À-pris", "Summa")
    for column_index, header in enumerate(headers, start=1):
        cell = summary.cell(row=4, column=column_index, value=header)
        cell.font = styles["th"]
        cell.fill = styles["fill_th"]
        cell.border = styles["border"]
    row = 5
    for line in offer["lines"]:
        values = (
            line["label"],
            line["stairwell_id"] or "—",
            float(line["quantity"]),
            line["unit"],
            float(line["unit_price"]),
            float(line["amount"]),
        )
        for column_index, value in enumerate(values, start=1):
            cell = summary.cell(row=row, column=column_index, value=value)
            cell.font = styles["b"]
            cell.border = styles["border"]
            if column_index in (5, 6):
                cell.number_format = money
        row += 1
    totals = offer["totals"]
    summary.cell(row=row, column=1, value="SUMMA STYCKPRISMODELL").font = styles["bb"]
    total_cell = summary.cell(row=row, column=6, value=float(totals["itemised_total"]))
    total_cell.font = styles["bb"]
    total_cell.number_format = money
    for column_index in range(1, 7):
        summary.cell(row=row, column=column_index).fill = styles["fill_sum"]
        summary.cell(row=row, column=column_index).border = styles["border"]
    row += 1
    summary.cell(
        row=row,
        column=1,
        value=f"REKOMMENDERAT FAST PRIS (risk {totals['risk_factor']})",
    ).font = styles["bb"]
    fixed_cell = summary.cell(row=row, column=6, value=float(totals["recommended_fixed_price"]))
    fixed_cell.font = styles["bb"]
    fixed_cell.number_format = money
    for column_index in range(1, 7):
        summary.cell(row=row, column=column_index).fill = styles["fill_sum"]
        summary.cell(row=row, column=column_index).border = styles["border"]
    row += 2
    for reservation in offer.get("reservations", []):
        summary.cell(
            row=row,
            column=1,
            value=(
                f"RESERVATION: {reservation['code']} × {reservation['quantity']} "
                f"saknar prisbokspost"
            ),
        ).font = styles["note"]
        row += 1

    lines_by_stairwell: dict[str, list[Mapping[str, Any]]] = {}
    for line in offer["lines"]:
        stairwell_id = line["stairwell_id"]
        if stairwell_id:
            lines_by_stairwell.setdefault(str(stairwell_id), []).append(line)
    for stairwell in service_quantities.get("stairwells", []):
        stairwell_id = str(stairwell["stairwell_id"])
        sheet = workbook.create_sheet(stairwell_id)
        for column, width in zip("ABCD", (36, 14, 14, 16), strict=False):
            sheet.column_dimensions[column].width = width
        sheet["A1"] = f"{stairwell_id} – provtryckningsmängder"
        sheet["A1"].font = styles["h1"]
        sheet["A1"].fill = styles["fill_h"]
        sheet.merge_cells("A1:D1")
        facts = (
            ("Antal lägenheter", stairwell["apartment_count"]),
            ("Schaktsträngar totalt", stairwell["string_count"]),
            ("Strängar att prova", stairwell["strings_to_test"]),
            ("Total stränglängd (m)", float(stairwell["total_length_m"])),
        )
        for offset, (label, value) in enumerate(facts):
            sheet.cell(row=3 + offset, column=1, value=label).font = styles["bb"]
            sheet.cell(row=3 + offset, column=2, value=value).font = styles["b"]
        row = 3 + len(facts) + 1
        for line in lines_by_stairwell.get(stairwell_id, []):
            sheet.cell(row=row, column=1, value=line["label"]).font = styles["b"]
            sheet.cell(row=row, column=2, value=float(line["quantity"])).font = styles["b"]
            amount_cell = sheet.cell(row=row, column=4, value=float(line["amount"]))
            amount_cell.font = styles["b"]
            amount_cell.number_format = money
            row += 1

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path
