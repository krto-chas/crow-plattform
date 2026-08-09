from __future__ import annotations

from decimal import Decimal
from pathlib import Path
from typing import Any

from crow_pressure_test import PressureTestKnowledge, TightnessClass

from .offer_workbook import _require_openpyxl, _styles

_EXAMPLE_PRESSURE_PA = 400
_TEST_ROWS = 25


def write_protocol_workbook(
    knowledge: PressureTestKnowledge,
    output_path: Path,
    project_title: str,
    required_class: TightnessClass = TightnessClass.C,
) -> Path:
    """Genererar täthetsprovningsprotokollet ur kunskapskällan.

    En enda sanning: läckagefaktorerna, ATC-mappningen och standardregistret
    kommer från lexikonet — protokollets kravtabell och q_max-formel kan
    aldrig divergera från kalkylens värden.
    """
    _require_openpyxl()
    import openpyxl

    styles = _styles()
    workbook = openpyxl.Workbook()

    requirements = workbook.active
    requirements.title = "Krav & standarder"
    for column, width in zip("ABCDE", (14, 18, 24, 18, 46), strict=False):
        requirements.column_dimensions[column].width = width
    requirements["A1"] = "TÄTHETSKLASSER – KRAV OCH STANDARDER"
    requirements["A1"].font = styles["h1"]
    requirements["A1"].fill = styles["fill_h"]
    requirements.merge_cells("A1:E1")
    headers = (
        "Täthetsklass",
        "Läckagefaktor c",
        f"Tillåtet vid {_EXAMPLE_PRESSURE_PA} Pa (l/s·m²)",
        "ATC (SS-EN 16798-3)",
        "Kommentar",
    )
    for column_index, header in enumerate(headers, start=1):
        cell = requirements.cell(row=3, column=column_index, value=header)
        cell.font = styles["th"]
        cell.fill = styles["fill_th"]
        cell.border = styles["border"]
    for offset, tightness_class in enumerate(TightnessClass):
        row = 4 + offset
        factor = knowledge.leakage_factor(tightness_class)
        example = knowledge.allowed_leakage_flow(
            tightness_class, _EXAMPLE_PRESSURE_PA, Decimal("1")
        )
        comment = "PROJEKTKRAV" if tightness_class is required_class else ""
        values: tuple[Any, ...] = (
            tightness_class.value,
            float(factor),
            float(example),
            knowledge.atc_class(tightness_class),
            comment,
        )
        for column_index, value in enumerate(values, start=1):
            cell = requirements.cell(row=row, column=column_index, value=value)
            cell.font = styles["bb"] if column_index == 1 else styles["b"]
            cell.border = styles["border"]
            if tightness_class is required_class:
                cell.fill = styles["fill_sum"]
    row = 4 + len(TightnessClass) + 1
    requirements.cell(
        row=row, column=1, value="STANDARDER SOM PROVNINGEN UTFÖRS MOT"
    ).font = styles["bb"]
    for offset, standard in enumerate(knowledge.standards()):
        requirements.cell(row=row + 1 + offset, column=1, value=standard.standard_id).font = (
            styles["bb"]
        )
        requirements.cell(row=row + 1 + offset, column=2, value=standard.title).font = styles[
            "note"
        ]

    protocol = workbook.create_sheet("Protokoll", 0)
    for column, width in zip("ABCDEFGH", (5, 32, 11, 11, 8, 14, 14, 22), strict=False):
        protocol.column_dimensions[column].width = width
    protocol["A1"] = f"PROTOKOLL – TÄTHETSPROVNING: {project_title}"
    protocol["A1"].font = styles["h1"]
    protocol["A1"].fill = styles["fill_h"]
    protocol.merge_cells("A1:H1")
    protocol["A2"] = (
        "q_max = c · |p|^0,65 · A. Klass och faktorer hämtas ur fliken "
        "Krav & standarder — samma källa som kalkylen."
    )
    protocol["A2"].font = styles["note"]
    header_fields = ("System", "Trapphus/del", "Datum", "Provad av", "Instrument/kalibrering")
    for offset, label in enumerate(header_fields):
        protocol.cell(row=4 + offset, column=1, value=label).font = styles["bb"]
        input_cell = protocol.cell(row=4 + offset, column=3)
        input_cell.fill = styles["fill_in"]
        input_cell.border = styles["border"]
        protocol.merge_cells(start_row=4 + offset, start_column=3, end_row=4 + offset, end_column=5)
    table_start = 4 + len(header_fields) + 1
    table_headers = (
        "Nr",
        "Provobjekt",
        "Yta A (m²)",
        "Tryck p (Pa)",
        "Klass",
        "q_max (l/s)",
        "Uppmätt (l/s)",
        "Resultat",
    )
    for column_index, header in enumerate(table_headers, start=1):
        cell = protocol.cell(row=table_start, column=column_index, value=header)
        cell.font = styles["th"]
        cell.fill = styles["fill_th"]
        cell.border = styles["border"]
    class_lookup = "'Krav & standarder'!$A$4:$B$7"
    for offset in range(_TEST_ROWS):
        row = table_start + 1 + offset
        protocol.cell(row=row, column=1, value=offset + 1).font = styles["note"]
        for column_index in (2, 3, 4, 5, 7):
            cell = protocol.cell(row=row, column=column_index)
            cell.fill = styles["fill_in"]
            cell.border = styles["border"]
        protocol.cell(
            row=row,
            column=6,
            value=(
                f'=IF(OR(C{row}="",D{row}="",E{row}=""),"",'
                f"VLOOKUP(E{row},{class_lookup},2,0)*ABS(D{row})^0.65*C{row})"
            ),
        ).border = styles["border"]
        protocol.cell(
            row=row,
            column=8,
            value=(
                f'=IF(OR(F{row}="",G{row}=""),"",'
                f'IF(G{row}<=F{row},"GODKÄND","EJ GODKÄND"))'
            ),
        ).border = styles["border"]

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path
